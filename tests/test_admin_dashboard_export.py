"""Admin dashboard xlsx export — prices + chart series, not an orders dump."""

from __future__ import annotations

import asyncio
import io
import os
import re
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from dotenv import load_dotenv
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.chart import LineChart

from app.admin_dashboard import XLSX_MEDIA_TYPE, normalize_range
from app.controllers import admin_controller as ac
from app.dashboard_xlsx import build_dashboard_xlsx


ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
os.environ.setdefault("STARTUP_SEED_MODE", "off")

CANONICAL_EXPORT = "/api/admin/dashboard/export"
_JS_EXPORT_RE = re.compile(r"""['"](/api/admin/dashboard/export[^'"?]*)""")


def _js_export_path() -> str:
    src = (ROOT / "public" / "js" / "api-client.js").read_text(encoding="utf-8")
    match = _JS_EXPORT_RE.search(src)
    assert match, "api-client.js missing dashboard export URL"
    return match.group(1)


def _book(body: bytes):
    assert body[:2] == b"PK"
    return load_workbook(io.BytesIO(body), data_only=False)


def _price_map(ws) -> dict:
    return {ws.cell(row, 1).value: ws.cell(row, 2) for row in range(2, ws.max_row + 1)}


def _day_cfg():
    return normalize_range(granularity="day", start="2026-08-01", end="2026-08-03")


def _orders():
    return [
        {
            "created_at": datetime(2026, 8, 1, 12, 0),
            "total_price": 10000,
            "status": "completed",
            "product_type": "戒指",
        },
        {
            "created_at": datetime(2026, 8, 2, 12, 0),
            "total_price": 20000,
            "status": "received",
            "product_type": "項墜",
        },
    ]


def _gold():
    return {"xau_per_gram": 4300.5, "xpt_per_gram": 1050, "xag_per_gram": 61.25}


def test_xlsx_has_price_and_chart_sheets_not_orders_dump():
    body, slug = build_dashboard_xlsx(_orders(), _day_cfg(), gold=_gold())
    wb = _book(body)
    assert wb.sheetnames == ["價格", "圖表"]
    price = wb["價格"]
    chart = wb["圖表"]
    assert [c.value for c in price[1]] == ["項目", "數值"]
    assert "訂單編號" not in [c.value for c in price[1]]
    assert "訂單編號" not in [c.value for c in chart[1]]
    assert slug == "2026-08-01_2026-08-03"


def test_price_sheet_kpis_and_metal_rates_are_numeric():
    cells = _price_map(_book(build_dashboard_xlsx(_orders(), _day_cfg(), gold=_gold())[0])["價格"])
    assert cells["區間"].data_type == "s"
    assert cells["完成營收"].value == 10000
    assert cells["完成營收"].data_type == "n"
    assert cells["完成營收"].number_format == "#,##0"
    assert cells["平均完成訂單"].value == 10000
    assert cells["平均完成訂單"].data_type == "n"
    assert cells["訂單總數"].value == 2
    assert cells["訂單總數"].data_type == "n"
    assert cells["黃金 XAU（NT$/公克）"].value == 4300.5
    assert cells["黃金 XAU（NT$/公克）"].data_type == "n"
    assert cells["黃金 XAU（NT$/公克）"].number_format == "#,##0.00"
    assert cells["白金 XPT（NT$/公克）"].value == 1050
    assert cells["白金 XPT（NT$/公克）"].data_type == "n"
    assert cells["白銀 XAG（NT$/公克）"].value == 61.25
    assert "$" not in str(cells["完成營收"].value)


def test_chart_sheet_dates_and_series_match_graph():
    ws = _book(build_dashboard_xlsx(_orders(), _day_cfg(), gold=None)[0])["圖表"]
    headers = [c.value for c in ws[1]]
    assert headers[:7] == [
        "日期",
        "區間",
        "標籤",
        "訂單總額",
        "完成營收",
        "訂單數量",
        "完成訂單數",
    ]
    row2 = {headers[i]: ws.cell(2, i + 1) for i in range(len(headers))}
    assert isinstance(row2["日期"].value, datetime)
    assert row2["日期"].value == datetime(2026, 8, 1)
    assert row2["日期"].data_type == "d"
    assert row2["日期"].number_format == "yyyy-mm-dd"
    assert row2["區間"].value == "2026-08-01"
    assert row2["訂單總額"].value == 10000
    assert row2["訂單總額"].data_type == "n"
    assert row2["完成營收"].value == 10000
    assert row2["訂單數量"].value == 1
    assert row2["完成訂單數"].value == 1
    row3 = {headers[i]: ws.cell(3, i + 1) for i in range(len(headers))}
    assert row3["日期"].value == datetime(2026, 8, 2)
    assert row3["訂單總額"].value == 20000
    assert row3["完成營收"].value == 0
    assert row3["累計訂單總額"].value == 30000
    assert row3["累計完成營收"].value == 10000


def _chart_title(chart) -> str:
    return chart.title.tx.rich.p[0].r[0].t


def test_chart_sheet_embeds_revenue_trend_line_chart():
    ws = _book(build_dashboard_xlsx(_orders(), _day_cfg(), gold=None)[0])["圖表"]
    assert len(ws._charts) >= 1
    chart = ws._charts[0]
    assert isinstance(chart, LineChart)
    assert _chart_title(chart) == "營收趨勢"
    assert len(chart.series) >= 2
    title_refs = [s.tx.strRef.f.replace("$", "") for s in chart.series[:2]]
    assert any(r.endswith("!D1") for r in title_refs)
    assert any(r.endswith("!E1") for r in title_refs)


def test_admin_html_dashboard_export_button_and_cache_bust():
    for name, version in (("admin1.html", "26"), ("admin.html", "24")):
        html = (ROOT / name).read_text(encoding="utf-8")
        assert 'id="dashExportBtn"' in html
        assert "匯出 Excel" in html
        assert f"admin.js?v={version}" in html
    src = (ROOT / "public" / "js" / "admin.js").read_text(encoding="utf-8")
    assert "Date.now()" in src
    assert "dashboardExportUrl" in src


def test_dashboard_js_downloads_canonical_export():
    assert _js_export_path() == CANONICAL_EXPORT
    src = (ROOT / "public" / "js" / "api-client.js").read_text(encoding="utf-8")
    assert "/api/admin/dashboard/export" in src
    assert "/api/admin/dashboard/export.xlsx" not in src


class _Cur:
    def execute(self, sql, params=None):
        return None

    def fetchall(self):
        return []

    def fetchone(self):
        return None

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


class _Conn:
    def __init__(self, cur):
        self._cur = cur

    def cursor(self):
        return self._cur

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


def _admin_api_client():
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    app = FastAPI()
    app.include_router(ac.router, prefix="/api")
    return TestClient(app)


def _full_app_client():
    from fastapi.testclient import TestClient

    from app import create_app

    return TestClient(create_app())


def _patch_export(monkeypatch, gold=None):
    monkeypatch.setattr(ac, "_require_admin", lambda _req: "admin-1")
    monkeypatch.setattr(ac, "_fetch_orders", lambda *a, **k: _orders())
    monkeypatch.setattr(ac, "_fetch_dashboard_gold", lambda _cur: gold if gold is not None else _gold())
    monkeypatch.setattr(ac, "get_connection", lambda: _Conn(_Cur()))


def test_export_endpoint_sets_xlsx_filename(monkeypatch):
    _patch_export(monkeypatch)
    resp = asyncio.run(
        ac.dashboard_export(
            MagicMock(),
            granularity="day",
            period=None,
            start="2026-08-01",
            end="2026-08-03",
        )
    )
    assert resp.media_type == XLSX_MEDIA_TYPE
    disp = resp.headers["content-disposition"]
    assert disp.startswith("attachment;")
    assert "dashboard-" in disp and disp.endswith('.xlsx"')
    body = resp.body if isinstance(resp.body, (bytes, bytearray)) else bytes(resp.body)
    cells = _price_map(_book(body)["價格"])
    assert cells["完成營收"].value == 10000


def test_export_endpoint_requires_admin(monkeypatch):
    def deny(_req):
        raise HTTPException(status_code=401, detail="not signed in")

    monkeypatch.setattr(ac, "_require_admin", deny)
    with pytest.raises(HTTPException) as exc:
        asyncio.run(ac.dashboard_export(MagicMock(), None, None, None, None))
    assert exc.value.status_code == 401


def test_export_http_get_returns_200_for_admin(monkeypatch):
    _patch_export(monkeypatch)
    client = _admin_api_client()
    resp = client.get(CANONICAL_EXPORT)
    assert resp.status_code == 200, resp.text[:180]
    assert resp.content[:2] == b"PK"
    assert "spreadsheetml" in (resp.headers.get("content-type") or "")
    assert (resp.headers.get("content-disposition") or "").endswith('.xlsx"')


def test_export_http_get_without_admin_is_401_not_404():
    client = _admin_api_client()
    resp = client.get(CANONICAL_EXPORT)
    assert resp.status_code == 401, (resp.status_code, resp.text[:180])
    assert resp.json().get("detail") != "Not Found"


def test_full_app_export_get_returns_200_for_admin(monkeypatch):
    _patch_export(monkeypatch)
    with _full_app_client() as client:
        resp = client.get(CANONICAL_EXPORT)
        assert resp.status_code == 200, (resp.status_code, resp.text[:180])
        assert resp.content[:2] == b"PK"
        ctype = resp.headers.get("content-type") or ""
        assert "spreadsheetml" in ctype
        disp = resp.headers.get("content-disposition") or ""
        assert disp.endswith('.xlsx"')
        assert "dashboard-" in disp


def test_full_app_export_get_without_admin_is_401_not_404():
    with _full_app_client() as client:
        resp = client.get(CANONICAL_EXPORT)
        assert resp.status_code == 401, (resp.status_code, resp.text[:180])
        assert resp.json() == {"detail": "not signed in"}
