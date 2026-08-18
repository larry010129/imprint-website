"""Admin orders xlsx export — Excel cell types, filters, auth."""

from __future__ import annotations

import asyncio
import inspect
import io
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from dotenv import load_dotenv
from fastapi import HTTPException
from openpyxl import load_workbook

from app.admin_dashboard import (
    ADMIN_ORDERS_CSV_HEADERS,
    XLSX_MEDIA_TYPE,
    build_admin_orders_xlsx,
)
from app.controllers import admin_controller as ac


ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
os.environ.setdefault("STARTUP_SEED_MODE", "off")

HEADERS = list(ADMIN_ORDERS_CSV_HEADERS)
CANONICAL_EXPORT = "/api/admin/orders/export"
EXPORT_PATHS = (
    CANONICAL_EXPORT,
    "/api/admin/orders/export.xlsx",
    "/api/admin/orders/export.csv",
)
_JS_EXPORT_RE = re.compile(r"""['"](/api/admin/orders/export[^'"?]*)""")


def _js_export_path() -> str:
    src = (ROOT / "public" / "js" / "admin-orders.js").read_text(encoding="utf-8")
    match = _JS_EXPORT_RE.search(src)
    assert match, "admin-orders.js missing export download URL"
    return match.group(1)


def _full_app_download_paths() -> list[str]:
    return list(dict.fromkeys([CANONICAL_EXPORT, _js_export_path()]))


def _order(**overrides):
    row = {
        "order_number": "IMP-1001",
        "created_at": datetime(2026, 8, 13, 3, 4),
        "status": "received",
        "customer_name": "王小明",
        "customer_email": "ming@example.com",
        "customer_phone": "0912345678",
        "summary": "經典系列 · 戒指",
        "summary_zh": "戒指 18K",
        "total_price": 40000,
        "shipping_postal": "106",
        "shipping_city": "台北市大安區",
        "shipping_address": "信義路一段1號",
        "cancel_reason": None,
    }
    row.update(overrides)
    return row


def _sheet(body: bytes):
    assert body[:2] == b"PK"
    return load_workbook(io.BytesIO(body), data_only=False).active


def _cells(ws, row=2) -> dict:
    return {HEADERS[i]: ws.cell(row, i + 1) for i in range(len(HEADERS))}


def test_xlsx_has_headers():
    ws = _sheet(build_admin_orders_xlsx([]))
    assert [c.value for c in ws[1]] == HEADERS
    assert ws.title == "訂單"


def test_xlsx_row_columns_and_zh_status():
    cells = _cells(_sheet(build_admin_orders_xlsx([_order()])))
    assert cells["訂單編號"].value == "IMP-1001"
    assert cells["訂單編號"].data_type == "s"
    assert cells["訂單編號"].number_format == "@"
    assert isinstance(cells["日期"].value, datetime)
    assert cells["日期"].value == datetime(2026, 8, 13, 3, 4)
    assert cells["日期"].data_type == "d"
    assert cells["日期"].number_format == "yyyy-mm-dd hh:mm"
    assert cells["狀態"].value == "已收到申請"
    assert cells["客戶姓名"].value == "王小明"
    assert cells["電子郵件"].value == "ming@example.com"
    assert cells["電話"].value == "0912345678"
    assert cells["電話"].data_type == "s"
    assert cells["電話"].number_format == "@"
    assert cells["品項"].value == "經典系列 · 戒指"
    assert cells["訂金"].value == 20000
    assert cells["總計"].value == 40000
    assert isinstance(cells["訂金"].value, int)
    assert isinstance(cells["總計"].value, int)
    assert cells["訂金"].data_type == "n"
    assert cells["總計"].data_type == "n"
    assert cells["訂金"].number_format == "#,##0"
    assert cells["總計"].number_format == "#,##0"
    assert cells["收件地址"].value == "106 台北市大安區 信義路一段1號"
    assert cells["取消原因"].value in ("", None)
    assert "$" not in str(cells["訂金"].value)
    assert "$" not in str(cells["總計"].value)
    assert "\t" not in str(cells["訂單編號"].value)
    assert "\t" not in str(cells["電話"].value)


def test_xlsx_phone_keeps_leading_zero():
    cells = _cells(
        _sheet(build_admin_orders_xlsx([_order(customer_phone=" 0900000000    ")]))
    )
    assert cells["電話"].value == "0900000000"
    assert cells["電話"].value.startswith("0")
    assert cells["電話"].data_type == "s"
    assert cells["電話"].number_format == "@"
    assert isinstance(cells["日期"].value, datetime)


def test_xlsx_money_is_numeric_half_deposit():
    cells = _cells(_sheet(build_admin_orders_xlsx([_order(total_price=45021)])))
    assert cells["總計"].value == 45021
    assert cells["訂金"].value == 22510
    assert isinstance(cells["訂金"].value, int)
    assert isinstance(cells["總計"].value, int)
    assert cells["訂金"].data_type == "n"
    assert cells["總計"].data_type == "n"


def test_xlsx_taiwan_tz_and_pickup_status():
    utc = datetime(2026, 8, 13, 14, 4, tzinfo=timezone.utc)
    cells = _cells(
        _sheet(
            build_admin_orders_xlsx(
                [_order(created_at=utc, status="shipped", fulfillment_method="pickup")]
            )
        )
    )
    assert cells["日期"].value == datetime(2026, 8, 13, 22, 4)
    assert cells["日期"].data_type == "d"
    assert cells["狀態"].value == "可取貨"


def test_xlsx_neutralizes_formulas():
    cells = _cells(
        _sheet(
            build_admin_orders_xlsx(
                [
                    _order(
                        customer_name="王, 小明",
                        customer_email="=1+1",
                        shipping_postal="",
                        shipping_city="",
                        shipping_address="=cmd",
                        cancel_reason="客戶要求取消",
                    )
                ]
            )
        )
    )
    assert cells["客戶姓名"].value == "王, 小明"
    assert cells["電子郵件"].value == "'=1+1"
    assert cells["電子郵件"].data_type == "s"
    assert str(cells["收件地址"].value).startswith("'=")
    assert cells["取消原因"].value == "客戶要求取消"


def test_admin1_and_admin_html_have_export_button():
    for name in ("admin1.html", "admin.html"):
        html = (ROOT / name).read_text(encoding="utf-8")
        assert 'id="btnExportOrdersCsv"' in html
        assert "匯出 Excel" in html
        assert "admin-orders.js?v=27" in html


def test_admin_orders_js_downloads_canonical_export():
    src = (ROOT / "public" / "js" / "admin-orders.js").read_text(encoding="utf-8")
    assert "/api/admin/orders/export?" in src
    assert "/api/admin/orders/export.xlsx" not in src
    assert _js_export_path() == CANONICAL_EXPORT
    assert "btnExportOrdersCsv" in src
    assert "Date.now()" in src


def test_export_search_sql_skips_config_json_cast():
    src = inspect.getsource(ac._admin_orders_search_from) + inspect.getsource(
        ac._fetch_admin_orders
    )
    assert "config_json::text" not in src
    assert "oi.summary_zh ilike" in src
    assert "o.order_number ilike" in src


class _Cur:
    def __init__(self, rows):
        self.rows = rows
        self.executed = []

    def execute(self, sql, params=None):
        self.executed.append((sql, params))

    def fetchall(self):
        return list(self.rows)

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


class _Conn:
    def __init__(self, cur):
        self._cur = cur

    def cursor(self):
        return self._cur

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


def test_export_endpoint_filters_by_q_and_sets_filename(monkeypatch):
    monkeypatch.setattr(ac, "_require_admin", lambda _req: "admin-1")
    monkeypatch.setattr(ac, "attach_order_display", lambda _cur, rows: None)
    cur = _Cur([_order(cancel_reason="重複訂單")])
    monkeypatch.setattr(ac, "get_connection", lambda: _Conn(cur))

    resp = asyncio.run(ac.orders_export(MagicMock(), q="IMP-1001"))
    assert "spreadsheetml" in (resp.media_type or "")
    assert resp.media_type == XLSX_MEDIA_TYPE
    disp = resp.headers["content-disposition"]
    assert disp.startswith("attachment;")
    assert "orders-" in disp and disp.endswith('.xlsx"')
    body = resp.body if isinstance(resp.body, (bytes, bytearray)) else bytes(resp.body)
    cells = _cells(_sheet(body))
    assert cells["訂單編號"].value == "IMP-1001"
    assert cells["取消原因"].value == "重複訂單"
    assert cur.executed
    sql, params = cur.executed[0]
    assert "ilike" in sql.lower()
    assert any(p == "%IMP-1001%" for p in (params or ()))


def test_export_endpoint_requires_admin(monkeypatch):
    def deny(_req):
        raise HTTPException(status_code=401, detail="not signed in")

    monkeypatch.setattr(ac, "_require_admin", deny)
    with pytest.raises(HTTPException) as exc:
        asyncio.run(ac.orders_export(MagicMock(), q=None))
    assert exc.value.status_code == 401


def _admin_api_client():
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    app = FastAPI()
    app.include_router(ac.router, prefix="/api")
    return TestClient(app)


def test_export_http_get_returns_200_for_admin(monkeypatch):
    monkeypatch.setattr(ac, "_require_admin", lambda _req: "admin-1")
    monkeypatch.setattr(ac, "attach_order_display", lambda _cur, rows: None)
    monkeypatch.setattr(ac, "get_connection", lambda: _Conn(_Cur([])))
    client = _admin_api_client()
    for path in EXPORT_PATHS:
        resp = client.get(path)
        assert resp.status_code == 200, (path, resp.status_code, resp.text[:180])
        assert resp.content[:2] == b"PK"
        ctype = resp.headers.get("content-type") or ""
        assert "spreadsheetml" in ctype
        disp = resp.headers.get("content-disposition") or ""
        assert disp.endswith('.xlsx"')


def test_export_http_get_without_admin_is_401_not_404():
    client = _admin_api_client()
    for path in EXPORT_PATHS:
        resp = client.get(path)
        assert resp.status_code == 401, (path, resp.status_code, resp.text[:180])
        assert resp.json().get("detail") != "Not Found"


def _full_app_client():
    from fastapi.testclient import TestClient

    from app import create_app

    return TestClient(create_app())


def test_full_app_export_get_returns_200_for_admin(monkeypatch):
    monkeypatch.setattr(ac, "_require_admin", lambda _req: "admin-1")
    monkeypatch.setattr(ac, "attach_order_display", lambda _cur, rows: None)
    monkeypatch.setattr(ac, "get_connection", lambda: _Conn(_Cur([])))
    with _full_app_client() as client:
        for path in _full_app_download_paths():
            resp = client.get(path)
            assert resp.status_code == 200, (path, resp.status_code, resp.text[:180])
            assert resp.content[:2] == b"PK"
            ctype = resp.headers.get("content-type") or ""
            assert "spreadsheetml" in ctype
            disp = resp.headers.get("content-disposition") or ""
            assert disp.endswith('.xlsx"')
            assert "orders-" in disp


def test_full_app_export_get_without_admin_is_401_not_404():
    with _full_app_client() as client:
        for path in _full_app_download_paths():
            resp = client.get(path)
            assert resp.status_code == 401, (path, resp.status_code, resp.text[:180])
            assert resp.json() == {"detail": "not signed in"}
            assert resp.json().get("detail") != "Not Found"
