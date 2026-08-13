"""Dashboard Excel export — KPI prices + chart time series."""

from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

from app.admin_dashboard import build_dashboard_payload

_MONEY = "#,##0"
_GOLD = "#,##0.00"
_DATE = "yyyy-mm-dd"
_GRAN_ZH = {"day": "日", "week": "週", "month": "月"}
_CHART_HEADERS = (
    "日期",
    "區間",
    "標籤",
    "訂單總額",
    "完成營收",
    "訂單數量",
    "完成訂單數",
    "累計訂單總額",
    "累計完成營收",
    "累計訂單數量",
    "累計完成訂單數",
)
_CHART_FMTS = (
    _DATE,
    "@",
    "@",
    _MONEY,
    _MONEY,
    _MONEY,
    _MONEY,
    _MONEY,
    _MONEY,
    _MONEY,
    _MONEY,
)


def _as_int(value) -> int:
    return round(float(value or 0))


def _as_money2(value):
    if value is None or value == "":
        return None
    return round(float(value), 2)


def _append_formatted(ws, values: list, formats: tuple[str, ...]) -> None:
    ws.append(values)
    row = ws.max_row
    for col, fmt in enumerate(formats, 1):
        ws.cell(row=row, column=col).number_format = fmt


def _bucket_date(key: str, gran: str) -> datetime | None:
    try:
        if gran == "day":
            return datetime.strptime(key, "%Y-%m-%d")
        if gran == "month":
            return datetime.strptime(key, "%Y-%m")
        if gran == "week":
            year, week = key.split("-W")
            return datetime.combine(
                date.fromisocalendar(int(year), int(week), 1), datetime.min.time()
            )
    except (TypeError, ValueError):
        return None
    return None


def _price_rows(payload: dict, gold: dict | None) -> list[tuple]:
    gran = payload["granularity"]
    rows = [
        ("區間", payload["periodLabel"], "@"),
        ("單位", _GRAN_ZH.get(gran, gran), "@"),
        ("完成營收", _as_int(payload["totalRevenue"]), _MONEY),
        ("平均完成訂單", _as_int(payload["averageSale"]), _MONEY),
        ("訂單總數", int(payload["periodOrderCount"] or 0), _MONEY),
    ]
    if not gold:
        return rows
    rows.extend(
        [
            ("黃金 XAU（NT$/公克）", _as_money2(gold.get("xau_per_gram")), _GOLD),
            ("白金 XPT（NT$/公克）", _as_money2(gold.get("xpt_per_gram")), _GOLD),
            ("白銀 XAG（NT$/公克）", _as_money2(gold.get("xag_per_gram")), _GOLD),
        ]
    )
    return rows


def _write_price_sheet(ws, payload: dict, gold: dict | None) -> None:
    ws.title = "價格"
    ws.append(["項目", "數值"])
    for label, value, fmt in _price_rows(payload, gold):
        _append_formatted(ws, [label, value], ("@", fmt))


def _chart_values(item: dict, gran: str, run: dict) -> list:
    total = _as_int(item.get("orderTotal"))
    revenue = _as_int(item.get("revenue"))
    count = int(item.get("orderCount") or 0)
    done = int(item.get("completedOrders") or 0)
    run["total"] += total
    run["revenue"] += revenue
    run["count"] += count
    run["done"] += done
    return [
        _bucket_date(str(item.get("month") or ""), gran),
        item.get("month") or "",
        item.get("label") or "",
        total,
        revenue,
        count,
        done,
        run["total"],
        run["revenue"],
        run["count"],
        run["done"],
    ]


def _add_trend_chart(ws) -> None:
    last = ws.max_row
    if last < 2:
        return
    chart = LineChart()
    chart.title = "營收趨勢"
    chart.style = 10
    chart.y_axis.numFmt = _MONEY
    chart.y_axis.scaling.min = 0
    chart.legend.position = "b"
    chart.height = 10
    chart.width = 18
    data = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=last)
    cats = Reference(ws, min_col=3, min_row=2, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "M2")


def _write_chart_sheet(ws, payload: dict) -> None:
    ws.append(list(_CHART_HEADERS))
    gran = payload["granularity"]
    run = {"total": 0, "revenue": 0, "count": 0, "done": 0}
    for item in payload.get("monthlyTrend") or []:
        _append_formatted(ws, _chart_values(item, gran, run), _CHART_FMTS)
    _add_trend_chart(ws)


def build_dashboard_xlsx(
    orders: list[dict], cfg: dict, gold: dict | None = None
) -> tuple[bytes, str]:
    payload = build_dashboard_payload(orders, cfg)
    wb = Workbook()
    _write_price_sheet(wb.active, payload, gold)
    _write_chart_sheet(wb.create_sheet("圖表"), payload)
    slug = payload["period"] or f"{payload['start']}_{payload['end']}" or "all"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), slug
