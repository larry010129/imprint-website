"""Admin dashboard aggregates — day/week/month ranges and Excel export."""

from __future__ import annotations

import io
import re
from datetime import date, datetime, time, timedelta, timezone

from openpyxl import Workbook

from app.orders import order_status_label

GRANULARITIES = ("day", "week", "month")
MAX_DAY_SPAN = 90
DAY_TREND_DEFAULT = 30
WEEK_TREND_COUNT = 12
MONTH_TREND_COUNT = 12
COMPLETE_STATUSES = {"completed", "shipped"}
# Safety cap so a bad range never hydrates the whole orders table.
DASHBOARD_ORDER_HARD_LIMIT = 5000

_WEEK_RE = re.compile(r"^(\d{4})-W(\d{2})$")
_TZ_TAIPEI = timezone(timedelta(hours=8))


def _now_local() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=8)


def _parse_date(value: str | None) -> date | None:
    try:
        return datetime.strptime(value or "", "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _parse_month(value: str | None) -> str | None:
    try:
        datetime.strptime(value or "", "%Y-%m")
        return value
    except (TypeError, ValueError):
        return None


def _parse_week(value: str | None) -> str | None:
    match = _WEEK_RE.match(value or "")
    if not match:
        return None
    year, week = int(match.group(1)), int(match.group(2))
    try:
        date.fromisocalendar(year, week, 1)
        return f"{year}-W{week:02d}"
    except ValueError:
        return None


def _week_start(week_str: str) -> date:
    year, week = map(int, week_str.split("-W"))
    return date.fromisocalendar(year, week, 1)


def _week_label(week_str: str) -> str:
    start = _week_start(week_str)
    end = start + timedelta(days=6)
    return f"{start.month}/{start.day}–{end.month}/{end.day}"


def _month_bucket_keys(now_local: datetime, count: int = MONTH_TREND_COUNT) -> list[str]:
    keys: list[str] = []
    year, month_num = now_local.year, now_local.month
    for offset in range(count - 1, -1, -1):
        index = year * 12 + (month_num - 1) - offset
        y, m = divmod(index, 12)
        keys.append(f"{y:04d}-{m + 1:02d}")
    return keys


def _week_bucket_keys(now_local: datetime, count: int = WEEK_TREND_COUNT) -> list[str]:
    today = now_local.date()
    iso = today.isocalendar()
    current = date.fromisocalendar(iso.year, iso.week, 1)
    keys: list[str] = []
    for offset in range(count - 1, -1, -1):
        week_start = current - timedelta(weeks=offset)
        iso_week = week_start.isocalendar()
        keys.append(f"{iso_week.year}-W{iso_week.week:02d}")
    return keys


def _day_bucket_keys(start_date: date, end_date: date) -> list[str]:
    keys: list[str] = []
    cursor = start_date
    while cursor <= end_date:
        keys.append(cursor.strftime("%Y-%m-%d"))
        cursor += timedelta(days=1)
    return keys


def _order_week_key(created_at: datetime | None) -> str:
    if not created_at:
        return ""
    local = created_at + timedelta(hours=8) if created_at.tzinfo else created_at
    iso = local.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def _order_day_key(created_at: datetime | None) -> str:
    if not created_at:
        return ""
    local = created_at + timedelta(hours=8) if created_at.tzinfo else created_at
    return local.strftime("%Y-%m-%d")


def _order_month_key(created_at: datetime | None) -> str:
    if not created_at:
        return ""
    local = created_at + timedelta(hours=8) if created_at.tzinfo else created_at
    return local.strftime("%Y-%m")


def normalize_range(
    *,
    granularity: str | None = None,
    period: str | None = None,
    start: str | None = None,
    end: str | None = None,
) -> dict:
    now_local = _now_local()
    today = now_local.date()
    gran = (granularity or "month").strip().lower()
    if gran not in GRANULARITIES:
        gran = "month"

    month_keys = _month_bucket_keys(now_local)
    week_keys = _week_bucket_keys(now_local)
    month_options = [{"value": key, "label": key} for key in reversed(month_keys)]
    week_options = [{"value": key, "label": _week_label(key)} for key in reversed(week_keys)]

    if gran == "month":
        selected = _parse_month(period) or now_local.strftime("%Y-%m")
        bucket_keys = month_keys
        period_label = selected
        range_start = range_end = None
    elif gran == "week":
        selected = _parse_week(period) or _order_week_key(now_local)
        bucket_keys = week_keys
        period_label = _week_label(selected)
        range_start = range_end = None
    else:
        end_date = _parse_date(end) or today
        start_date = _parse_date(start) or (end_date - timedelta(days=DAY_TREND_DEFAULT - 1))
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        if (end_date - start_date).days > MAX_DAY_SPAN - 1:
            start_date = end_date - timedelta(days=MAX_DAY_SPAN - 1)
        bucket_keys = _day_bucket_keys(start_date, end_date)
        selected = None
        period_label = f"{start_date.strftime('%Y-%m-%d')} – {end_date.strftime('%Y-%m-%d')}"
        range_start = start_date.strftime("%Y-%m-%d")
        range_end = end_date.strftime("%Y-%m-%d")

    return {
        "granularity": gran,
        "period": selected or "",
        "start": range_start or (today - timedelta(days=DAY_TREND_DEFAULT - 1)).strftime("%Y-%m-%d"),
        "end": range_end or today.strftime("%Y-%m-%d"),
        "periodLabel": period_label,
        "monthOptions": month_options,
        "weekOptions": week_options,
        "bucketKeys": bucket_keys,
    }


def dashboard_fetch_bounds(cfg: dict) -> tuple[datetime, datetime]:
    """UTC [since, until) covering cfg bucketKeys (local +8 calendar).

    Dashboard trend needs the full bucket window, not only the selected period.
    Callers must filter SQL with these bounds — never hydrate all orders.
    """
    keys = list(cfg.get("bucketKeys") or [])
    gran = cfg.get("granularity") or "month"

    if gran == "day" and keys:
        start_d = date.fromisoformat(keys[0])
        end_exclusive = date.fromisoformat(keys[-1]) + timedelta(days=1)
    elif gran == "week" and keys:
        start_d = _week_start(keys[0])
        end_exclusive = _week_start(keys[-1]) + timedelta(days=7)
    elif gran == "month" and keys:
        y, m = map(int, keys[0].split("-"))
        start_d = date(y, m, 1)
        y2, m2 = map(int, keys[-1].split("-"))
        if m2 == 12:
            end_exclusive = date(y2 + 1, 1, 1)
        else:
            end_exclusive = date(y2, m2 + 1, 1)
    else:
        start_d = _parse_date(cfg.get("start")) or (_now_local().date() - timedelta(days=DAY_TREND_DEFAULT - 1))
        end_d = _parse_date(cfg.get("end")) or _now_local().date()
        if start_d > end_d:
            start_d, end_d = end_d, start_d
        end_exclusive = end_d + timedelta(days=1)

    since = datetime.combine(start_d, time.min, tzinfo=_TZ_TAIPEI).astimezone(timezone.utc)
    until = datetime.combine(end_exclusive, time.min, tzinfo=_TZ_TAIPEI).astimezone(timezone.utc)
    return since, until


def _order_in_period(order: dict, cfg: dict) -> bool:
    created_at = order.get("created_at")
    gran = cfg["granularity"]
    if gran == "day":
        key = _order_day_key(created_at)
        return bool(key) and cfg["start"] <= key <= cfg["end"]
    if gran == "week":
        return _order_week_key(created_at) == cfg["period"]
    return _order_month_key(created_at) == cfg["period"]


def _trend_label(key: str, granularity: str) -> str:
    if granularity == "month" and len(key) >= 7:
        return f"{int(key[5:7])}月"
    if granularity == "week":
        return _week_label(key)
    if granularity == "day" and len(key) >= 10:
        return key[5:]
    return key


def _bucket_key(order: dict, granularity: str) -> str:
    created_at = order.get("created_at")
    if granularity == "day":
        return _order_day_key(created_at)
    if granularity == "week":
        return _order_week_key(created_at)
    return _order_month_key(created_at)


def build_dashboard_payload(orders: list[dict], cfg: dict) -> dict:
    bucket_keys = cfg["bucketKeys"]
    gran = cfg["granularity"]

    period_orders = [o for o in orders if _order_in_period(o, cfg)]
    period_completed = [o for o in period_orders if o.get("status") in COMPLETE_STATUSES]
    period_revenue = sum(float(o.get("total_price") or 0) for o in period_completed)
    period_order_count = len(period_orders)
    average_sale = period_revenue / len(period_completed) if period_completed else 0

    buckets: dict[str, dict] = {
        key: {"orderCount": 0, "completedOrders": 0, "orderTotal": 0.0, "revenue": 0.0}
        for key in bucket_keys
    }

    for order in orders:
        key = _bucket_key(order, gran)
        if key not in buckets:
            continue
        price = float(order.get("total_price") or 0)
        buckets[key]["orderCount"] += 1
        buckets[key]["orderTotal"] += price
        if order.get("status") in COMPLETE_STATUSES:
            buckets[key]["completedOrders"] += 1
            buckets[key]["revenue"] += price

    monthly_trend = [
        {
            "month": key,
            "label": _trend_label(key, gran),
            "orderCount": buckets[key]["orderCount"],
            "completedOrders": buckets[key]["completedOrders"],
            "orderTotal": buckets[key]["orderTotal"],
            "revenue": buckets[key]["revenue"],
        }
        for key in bucket_keys
    ]

    status_counts: dict[str, int] = {}
    for order in period_orders:
        status = order.get("status") or "unknown"
        status_counts[status] = status_counts.get(status, 0) + 1
    total_status = sum(status_counts.values()) or 1
    status_rows = [
        {
            "code": code,
            "count": count,
            "percent": round(100 * count / total_status, 1),
        }
        for code, count in sorted(status_counts.items(), key=lambda x: -x[1])
    ]

    product_counts: dict[str, int] = {}
    for order in period_orders:
        name = (order.get("product_type") or order.get("category") or "").strip() or "其他"
        product_counts[name] = product_counts.get(name, 0) + 1
    top_products = [
        {"name": name, "orders": count}
        for name, count in sorted(product_counts.items(), key=lambda x: -x[1])[:6]
    ]

    series_stats: dict[str, dict] = {}
    for order in period_orders:
        name = (order.get("series") or "").strip() or "未分類"
        if name not in series_stats:
            series_stats[name] = {"orders": 0, "revenue": 0.0}
        series_stats[name]["orders"] += 1
        if order.get("status") in COMPLETE_STATUSES:
            series_stats[name]["revenue"] += float(order.get("total_price") or 0)
    top_series = [
        {"name": name, "orders": vals["orders"], "revenue": vals["revenue"]}
        for name, vals in sorted(series_stats.items(), key=lambda x: -x[1]["orders"])[:6]
    ]

    recent_orders = sorted(orders, key=lambda o: o.get("created_at") or datetime.min, reverse=True)[:10]

    return {
        "granularity": gran,
        "period": cfg["period"],
        "start": cfg["start"],
        "end": cfg["end"],
        "periodLabel": cfg["periodLabel"],
        "monthOptions": cfg["monthOptions"],
        "weekOptions": cfg["weekOptions"],
        "periodRevenue": period_revenue,
        "periodOrderCount": period_order_count,
        "totalRevenue": period_revenue,
        "averageSale": round(average_sale, 2),
        "monthlyTrend": monthly_trend,
        "statusRows": status_rows,
        "topProducts": top_products,
        "topSeries": top_series,
        "totalOrders": period_order_count,
        "recentOrders": recent_orders,
    }


def _csv_safe(value) -> str:
    """Neutralize spreadsheet formula injection. A cell beginning with = + - @
    (or tab/CR) is treated as a formula by Excel/Sheets; customer_name and the
    product summary are attacker-controlled at checkout, so prefix a single
    quote to force text interpretation. See OWASP "CSV Injection"."""
    text = "" if value is None else str(value)
    if text and text[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + text
    return text


ADMIN_ORDERS_CSV_HEADERS = (
    "訂單編號",
    "日期",
    "狀態",
    "客戶姓名",
    "電子郵件",
    "電話",
    "品項",
    "訂金",
    "總計",
    "收件地址",
    "取消原因",
)


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_XLSX_COL_FORMATS = (
    "@",
    "yyyy-mm-dd hh:mm",
    "@",
    "@",
    "@",
    "@",
    "@",
    "#,##0",
    "#,##0",
    "@",
    "@",
)


def _xlsx_local_dt(value) -> datetime | None:
    if not value:
        return None
    dt = value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return None
    if not isinstance(dt, datetime):
        return None
    local = dt.astimezone(_TZ_TAIPEI) if dt.tzinfo else dt
    return local.replace(tzinfo=None, second=0, microsecond=0)


def _csv_shipping(order: dict) -> str:
    parts = [
        order.get("shipping_postal"),
        order.get("shipping_city"),
        order.get("shipping_address"),
    ]
    return " ".join(str(p).strip() for p in parts if p)


def _csv_items(order: dict) -> str:
    return (
        order.get("summary")
        or order.get("summary_zh")
        or order.get("product_name")
        or order.get("product_type")
        or ""
    )


def _csv_money(value):
    if value is None or value == "":
        return None
    return round(float(value))


def _csv_deposit(order: dict):
    total = order.get("total_price")
    if total is None or total == "":
        return None
    return round(float(total) * 0.5)


def _csv_status(order: dict) -> str:
    label = order.get("status_label")
    if label:
        return str(label)
    return order_status_label(order.get("status"), order.get("fulfillment_method"))


def _admin_order_xlsx_values(order: dict) -> list:
    phone = str(order.get("customer_phone") or "").strip()
    return [
        _csv_safe(str(order.get("order_number") or "").strip()),
        _xlsx_local_dt(order.get("created_at")),
        _csv_safe(_csv_status(order)),
        _csv_safe(order.get("customer_name") or ""),
        _csv_safe(order.get("customer_email") or ""),
        _csv_safe(phone),
        _csv_safe(_csv_items(order)),
        _csv_deposit(order),
        _csv_money(order.get("total_price")),
        _csv_safe(_csv_shipping(order)),
        _csv_safe(order.get("cancel_reason") or ""),
    ]


def _xlsx_apply_formats(ws, row_idx: int) -> None:
    for col, fmt in enumerate(_XLSX_COL_FORMATS, 1):
        ws.cell(row=row_idx, column=col).number_format = fmt


def build_admin_orders_xlsx(orders: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "訂單"
    ws.append(list(ADMIN_ORDERS_CSV_HEADERS))
    for order in orders:
        ws.append(_admin_order_xlsx_values(order))
        _xlsx_apply_formats(ws, ws.max_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
